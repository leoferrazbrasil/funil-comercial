from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


ROOT = Path.cwd()
DATE = "2026-07-23"
OUT_DIR = ROOT / f"prospeccao-ativa-{DATE}"
CSV_PATH = OUT_DIR / "leads-qualificados-2026-07-23.csv"
XLSX_PATH = OUT_DIR / "leads-qualificados-2026-07-23.xlsx"

STATUS_VALUES = [
    "LEAD_QUALIFICADO_AGUARDANDO_REDESIGN_INDIVIDUAL",
    "SEM_SITE_NAO_CONTA",
    "SITE_BOM_DESCARTADO",
    "NICHO_EXCLUIDO_NAO_CONTA",
    "CONTATO_INSUFICIENTE",
    "META_NAO_BATIDA",
]

with CSV_PATH.open("r", encoding="utf-8", newline="") as handle:
    rows = list(csv.reader(handle, delimiter=";"))

wb = Workbook()
ws = wb.active
ws.title = DATE

for row in rows:
    ws.append(row)

ws.freeze_panes = "K2"
ws.auto_filter.ref = f"A1:AF{ws.max_row}"

header_fill = PatternFill("solid", fgColor="111111")
critical_fill = PatternFill("solid", fgColor="FFD700")
white_font = Font(color="FFFFFF", bold=True)
black_font = Font(color="111111", bold=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

critical_cols = {9, 10, 21, 25, 26}
for cell in ws[1]:
    cell.fill = critical_fill if cell.column in critical_cols else header_fill
    cell.font = black_font if cell.column in critical_cols else white_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
    row[8].number_format = "@"
    row[20].value = None

widths = {
    "A": 12, "B": 8, "C": 22, "D": 24, "E": 28, "F": 14, "G": 14, "H": 38,
    "I": 24, "J": 44, "K": 42, "L": 48, "M": 56, "N": 68, "O": 48, "P": 44,
    "Q": 42, "R": 42, "S": 36, "T": 28, "U": 20, "V": 48, "W": 48, "X": 48,
    "Y": 34, "Z": 48, "AA": 44, "AB": 64, "AC": 56, "AD": 48, "AE": 52, "AF": 48,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.row_dimensions[1].height = 42
for idx in range(2, ws.max_row + 1):
    ws.row_dimensions[idx].height = 138

dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_VALUES) + '"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(f"Z2:Z{ws.max_row}")

status_colors = {
    "LEAD_QUALIFICADO_AGUARDANDO_REDESIGN_INDIVIDUAL": "D1EFD1",
    "SEM_SITE_NAO_CONTA": "E5E5E5",
    "SITE_BOM_DESCARTADO": "CCE5FF",
    "NICHO_EXCLUIDO_NAO_CONTA": "E0D1F4",
    "CONTATO_INSUFICIENTE": "FFF4B3",
    "META_NAO_BATIDA": "FFC6C6",
}
for status, color in status_colors.items():
    fill = PatternFill("solid", fgColor=color)
    ws.conditional_formatting.add(
        f"A2:AF{ws.max_row}",
        FormulaRule(formula=[f'$Z2="{status}"'], fill=fill),
    )

for row_idx in range(2, ws.max_row + 1):
    if row_idx % 2 == 0:
        for cell in ws[row_idx]:
            if cell.column != 26:
                cell.fill = PatternFill("solid", fgColor="FAFAFA")

summary = wb.create_sheet("Resumo")
summary.append(["item", "valor"])
summary.append(["data", DATE])
summary.append(["leads qualificados", ws.max_row - 1])
summary.append(["status principal", "LEAD_QUALIFICADO_AGUARDANDO_REDESIGN_INDIVIDUAL"])
summary.append(["URL curta gerada", "em branco por regra da automação"])
summary.append(["prints PNG", 90])
summary.append(["envio automático", "não realizado"])
summary.append(["redesign/build/publicação", "não realizado"])
for cell in summary[1]:
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center")
summary.column_dimensions["A"].width = 30
summary.column_dimensions["B"].width = 64

wb.save(XLSX_PATH)
print(XLSX_PATH)
