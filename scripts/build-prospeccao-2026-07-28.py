from pathlib import Path
import csv
import json
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


DATE = "2026-07-28"
ROOT = Path.cwd()
OUT_DIR = ROOT / "prospeccao" / DATE
PRINT_DIR = OUT_DIR / "prints"
CSV_PATH = OUT_DIR / f"prospeccao_{DATE}.csv"
XLSX_PATH = OUT_DIR / f"Prospeccao_Ativa_{DATE}.xlsx"
JSON_PATH = OUT_DIR / "qualification-capture.json"

COLUMNS = [
    "data",
    "rank",
    "nicho",
    "regiao",
    "nome",
    "nota",
    "avaliacoes",
    "contato",
    "WhatsApp/telefone normalizado",
    "site atual",
    "link Maps",
    "motivo da abordagem",
    "diagnostico do site atual",
    "conteudo real extraido",
    "servicos reais identificados",
    "prova social real",
    "identidade visual observada",
    "imagens/logo aproveitaveis",
    "slug sugerido",
    "status redesign",
    "URL curta gerada",
    "print Maps",
    "print site atual desktop",
    "print site atual mobile",
    "link wa.me preliminar",
    "status",
    "observacoes",
    "Bloco 1 (Link)",
    "Bloco 2 (Proposta)",
    "Toque 1 (2 dias)",
    "Toque 2 (5 dias)",
    "Toque 3 (10 dias)",
]

STATUS_VALUES = [
    "LEAD_QUALIFICADO_AGUARDANDO_REDESIGN_INDIVIDUAL",
    "SEM_SITE_NAO_CONTA",
    "SITE_BOM_DESCARTADO",
    "NICHO_EXCLUIDO_NAO_CONTA",
    "CONTATO_INSUFICIENTE",
    "META_NAO_BATIDA",
]

LEADS = [
    {
        "nicho": "Psicologa",
        "regiao": "Uberlandia - MG",
        "nome": "Laura Pereira Alves",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5534997240174",
        "site": "https://apsicoterapia.wixsite.com/psilaura",
        "contato": "Telefone/WhatsApp publico no site: +55 (34) 99724-0174.",
        "motivo": "Psicologa com CRP, atendimento online e cidade prioritaria, mas em dominio Wix e pagina simples.",
        "diagnostico": "Site Wix de pagina unica, com hierarquia visual basica, pouco destaque para agenda e ausencia de prova social organizada.",
        "conteudo": "Sou psicologa CRP 04/75579, formada pela Universidade Federal de Uberlandia; psicoterapia individual online e atividades em Uberlandia.",
        "servicos": "Psicoterapia individual online; oficinas, palestras e treinamentos para grupos e instituicoes.",
        "prova": "CRP 04/75579, formacao UFU e telefone publicados no site.",
        "identidade": "Wix minimalista com imagens leves e pouca marca proprietaria.",
        "imagens": "Imagem de perfil e elementos visuais simples podem orientar redesign.",
    },
    {
        "nicho": "Psicologa / Psicanalista",
        "regiao": "Belo Horizonte - MG",
        "nome": "Vanessa Ferraz",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5531993736983",
        "site": "https://vanessafda.wixsite.com/psicanalise",
        "contato": "Telefone publico no site: 31 99373-6983.",
        "motivo": "Psicologa e psicanalista em BH com CRP e experiencia publica, mas site Wix antigo.",
        "diagnostico": "Pagina Wix com galeria grande, texto longo e CTA pouco claro; falta estrutura de conversao para agendamento.",
        "conteudo": "Vanessa Ferraz, CRP-MG 04/44586, psicologa/psicanalista, perita do TJMG, formada pela FUMEC e com percurso em UFMG/FCMMG.",
        "servicos": "Psicanalise; psicologia clinica; assistencia tecnica/pericia.",
        "prova": "CRP, publicacoes e historico profissional publicados no site.",
        "identidade": "Visual artistico, frases de Freud e galeria pessoal, mas layout datado.",
        "imagens": "Fotos e galeria do consultorio/profissional podem ser reaproveitadas.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Belo Horizonte - MG",
        "nome": "Luiza Elias",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5531971442824",
        "site": "https://psicologaluizaelia.wixsite.com/website",
        "contato": "Telefone publico no site: 31 97144-2824; e-mail psicologa.luizaelias@gmail.com.",
        "motivo": "Atendimento de familia em BH com contato direto, mas pagina Wix curta e pouco persuasiva.",
        "diagnostico": "Site antigo, poucos blocos, ausencia de prova social e CTA sem destaque visual.",
        "conteudo": "Pagina informa terapia de familia e atendimento psicologico, endereco R. Padre Pedro Evangelista, 45, Belo Horizonte.",
        "servicos": "Terapia de familia; atendimento psicologico presencial.",
        "prova": "Telefone, e-mail e endereco publicados no site.",
        "identidade": "Wix simples com imagem pessoal e estrutura basica.",
        "imagens": "Foto/imagem do site pode ser base inicial.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Uberlandia - MG",
        "nome": "Leticia Santos",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5534991637775",
        "site": "https://lesantospsico.wixsite.com/leticiasantospsi",
        "contato": "WhatsApp publico no site: (34) 99163-7775; e-mail lesantospsico@gmail.com.",
        "motivo": "Psicologa em Uberlandia com CRP e especialidades claras, mas site Wix ainda pouco refinado.",
        "diagnostico": "Pagina longa com texto denso, design de template e CTA que pode ser mais direto para agenda.",
        "conteudo": "Leticia Santos, CRP 04/54636, formada pela UFU, atende criancas, adolescentes, adultos e idosos presencialmente em Uberlandia e online.",
        "servicos": "Psicanalise; depressao; ansiedade; panico; TEA; TDAH; dificuldades de aprendizagem.",
        "prova": "CRP 04/54636, formacao UFU, endereco e WhatsApp publicados.",
        "identidade": "Wix com texto institucional, fotos e poucos elementos de marca.",
        "imagens": "Fotos pessoais e elementos do site podem orientar redesign.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Belo Horizonte - MG",
        "nome": "Amanda Azevedo",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5535998049826",
        "site": "https://psiamandaazevedo.wixsite.com/amandaazevedo",
        "contato": "WhatsApp/e-mail publico: +55 (35) 99804-9826; psi.amandaazevedo@gmail.com.",
        "motivo": "Psicologa em Lourdes/BH com contato direto, mas presenca digital em Wix e baixa hierarquia comercial.",
        "diagnostico": "Site simples, com conteudo curto e pouco destaque para agendamento, diferenciais e prova social.",
        "conteudo": "Amanda Azevedo, atendimento em Belo Horizonte, endereco Rua dos Timbiras 1940, sala 1704, Lourdes.",
        "servicos": "Psicoterapia; atendimento presencial em BH.",
        "prova": "Telefone, e-mail, endereco, Instagram e LinkedIn publicados no site.",
        "identidade": "Wix leve, visual limpo e pouca personalidade de marca.",
        "imagens": "Fotos/elementos sociais podem ser aproveitados em pagina futura.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Uberlandia - MG",
        "nome": "Mirian Cristina da Silva Santos",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5534999224879",
        "site": "https://cristinassmirian.wixsite.com/psicomiriansantos/blank-1",
        "contato": "WhatsApp publico no site: (34) 99922-4879.",
        "motivo": "Psicologa TCC com CRP e historico profissional, mas site Wix antigo e visualmente datado.",
        "diagnostico": "Pagina com banner Wix, estrutura antiga, textos curriculares longos e CTA pouco forte.",
        "conteudo": "Mirian Santos, CRP 04/33254, psicoterapia clinica cognitivo-comportamental juvenil, psicologa clinica desde 2010.",
        "servicos": "Terapia individual; psicoterapia juvenil; TCC; avaliacao/reabilitacao neuropsicologica.",
        "prova": "CRP, formacao UFU e WhatsApp publicados no site.",
        "identidade": "Wix antigo com imagens simples e baixa percepcao de autoridade.",
        "imagens": "Foto profissional e elementos do site podem ser reaproveitados.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Uberlandia - MG",
        "nome": "Luciana Gomes",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5534991728418",
        "site": "https://lucianagomes125.wixsite.com/website",
        "contato": "Telefone publico no site: 34 99172-8418; e-mail lucianagpsico@gmail.com.",
        "motivo": "Psicologa em regiao prioritaria com site Wix antigo e contato direto.",
        "diagnostico": "Site de 2018 com estrutura simples, CTA discreto e pouca organizacao de beneficios para conversao.",
        "conteudo": "Psicologa Luciana Gomes, CRP 04/45227, endereco R. Duque de Caxias, 450 - Centro, Uberlandia.",
        "servicos": "Psicologia clinica; atendimento presencial.",
        "prova": "CRP, endereco, telefone, Facebook e Instagram publicados.",
        "identidade": "Wix antigo, poucos recursos visuais e linguagem institucional basica.",
        "imagens": "Imagens e redes indicadas podem orientar visual futuro.",
    },
    {
        "nicho": "Psicologa infantil",
        "regiao": "Nova Iguacu - RJ",
        "nome": "Veronica de Pontes Leandro",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5521975525914",
        "site": "https://vepontesesophia.wixsite.com/psiveronicapontes",
        "contato": "Telefone/WhatsApp publico no site: (21) 97552-5914.",
        "motivo": "Psicologa infantil em Nova Iguacu com WhatsApp e endereco, mas site Wix com layout basico.",
        "diagnostico": "Pagina com muitos botoes repetidos, hierarquia fraca e pouca prova social/autoridade.",
        "conteudo": "Veronica de Pontes Leandro, psicologa clinica para criancas e adolescentes, atendimento presencial e online no Edificio Top Commerce em Nova Iguacu.",
        "servicos": "Avaliacao psicologica; terapia individual para criancas e adolescentes; regulacao emocional.",
        "prova": "Telefone, endereco e servicos publicados no site.",
        "identidade": "Wix com galeria e botoes simples, baixa sofisticacao visual.",
        "imagens": "Galeria e imagens atuais podem orientar redesign.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Nova Iguacu - RJ",
        "nome": "Rachel Salsinha",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5521969691886",
        "site": "https://psicologarachelsal.wixsite.com/psiclogarachelsalsin",
        "contato": "WhatsApp publico no site: (21) 96969-1886.",
        "motivo": "Psicologa licenciada com CRP e WhatsApp, mas site Wix antigo e com URL improvisada.",
        "diagnostico": "Dominio Wix, texto generico e CTA simples; redesign pode destacar atendimento, CRP e agenda.",
        "conteudo": "Psicologa Rachel Salsinha, CRP 05/72004, Rua Getulio Vargas 121, Nova Iguacu/RJ.",
        "servicos": "Psicoterapia para trauma, ansiedade, vicio, sofrimento e mudancas de vida.",
        "prova": "CRP, WhatsApp e endereco publicados no site.",
        "identidade": "Wix com estrutura basica e imagem hero generica.",
        "imagens": "Fotos/imagens da pagina podem orientar uma identidade mais humana.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Tatui - SP",
        "nome": "Livia Longo",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5519998166078",
        "site": "https://liviaon5.wixsite.com/psicoterapia",
        "contato": "WhatsApp publico no site: (19) 99816-6078; e-mail liviaon@gmail.com.",
        "motivo": "Psicologa com CRP e WhatsApp no interior de SP, mas site Wix antigo.",
        "diagnostico": "Site com estrutura de template, informacoes repetidas e pouca clareza de conversao na primeira dobra.",
        "conteudo": "Livia Longo, psicologa CRP 06/168309, atendimento em Tatuí/SP, Centro Empresarial Barao de Tatui.",
        "servicos": "Psicoterapia; atendimento clinico presencial/online.",
        "prova": "CRP, telefone, endereco e e-mail publicados no site.",
        "identidade": "Wix com textos reflexivos, imagem e identidade pouco comercial.",
        "imagens": "Imagem de perfil/visual do site podem ser reaproveitados.",
    },
    {
        "nicho": "Psicologa / Terapeuta",
        "regiao": "Rio de Janeiro - RJ",
        "nome": "Amanda Figueiroa K. de Sant'Anna",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5521968702245",
        "site": "https://afksantanna.wixsite.com/amandafigueiroa",
        "contato": "Telefone publico no site: (21) 96870-2245; e-mail afksantanna@gmail.com.",
        "motivo": "Psicologa/terapeuta com telefone e conteudo, mas site Wix com baixa conversao.",
        "diagnostico": "Pagina simples, URL Wix, conteudo pouco escaneavel e CTA limitado.",
        "conteudo": "Pedagoga e psicologa graduada pela UERJ; psicoterapeuta transpessoal e regressao de memoria, atendimento jovens e adultos.",
        "servicos": "Psicoterapia transpessoal; terapia regressiva; atendimento jovens e adultos presencial e online.",
        "prova": "Formacao UERJ, telefone, e-mail e abordagem publicados no site.",
        "identidade": "Wix com simbolo/logo simples e texto conceitual longo.",
        "imagens": "Simbolo/logo e imagens atuais podem ser aproveitados.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Sao Jose dos Campos - SP",
        "nome": "Gabriella Gasafe",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5511933330383",
        "site": "https://psigabriellagasafe.wixsite.com/psicologia",
        "contato": "Telefone publico no site: (11) 93333-0383; e-mail psi.gabriellagasafe@gmail.com.",
        "motivo": "Psicologa com CRP e atendimento em Sao Jose dos Campos, mas ainda usa subdominio Wix.",
        "diagnostico": "Site tem conteudo real, porem em template Wix, com blocos longos e CTA que pode ganhar foco e prova social.",
        "conteudo": "Gabriella Gasafe, CRP 06/199953, Gestalt-terapia, atendimento online para adolescentes/adultos e presencial em SJC/Mogi.",
        "servicos": "Psicoterapia online; presencial; atendimento infantil; consultoria; orientacao profissional.",
        "prova": "CRP, enderecos, telefone e e-mail publicados no site.",
        "identidade": "Wix claro com foto e elementos delicados, mas pouco premium.",
        "imagens": "Foto, folha/elementos e textos do site podem orientar redesign.",
    },
    {
        "nicho": "Psicologa / Psicopedagoga",
        "regiao": "Sao Paulo - SP",
        "nome": "Cintia Velloso Malpelli",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5511999262926",
        "site": "https://www.psicocintiavelloso.com.br/",
        "contato": "WhatsApp publico no site: (11) 99926-2926.",
        "motivo": "Psicologa com dominio proprio e WhatsApp, mas pagina tem muitos blocos e comunicacao visual datada.",
        "diagnostico": "Site sobrecarregado, repeticao de logo e links, pouca hierarquia de oferta e CTA espalhado.",
        "conteudo": "Cintia Velloso Malpelli, CRP 06/137183, psicologa e psicopedagoga, psicoterapia individual e terapia de casal.",
        "servicos": "Psicoterapia individual; terapia de casal; psicopedagogia clinica.",
        "prova": "CRP, WhatsApp, artigos e canal publicados no site.",
        "identidade": "Identidade propria com logo, porem layout antigo e pouco responsivo.",
        "imagens": "Logo, fotos e materiais de livro/artigos podem ser aproveitados.",
    },
    {
        "nicho": "Psicanalista",
        "regiao": "Sao Paulo - SP",
        "nome": "Raissa Assuncao",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5511996068862",
        "site": "https://raissapsicanalista.wixsite.com/home",
        "contato": "Telefone/WhatsApp publico no site: (11) 99606-8862; e-mail psicanalistaraissaassuncao@gmail.com.",
        "motivo": "Psicanalista mulher em Sao Paulo com WhatsApp, mas site Wix de estrutura simples.",
        "diagnostico": "Pagina Wix curta, pouco foco em conversao, ausencia de prova social e URL improvisada.",
        "conteudo": "Raissa Assuncao, psicanalista pela Escola de Psicanalise de Sao Paulo e graduanda em Psicologia pela Mackenzie.",
        "servicos": "Psicanalise; atendimento clinico; producao academica sobre depressao e suicidio.",
        "prova": "Telefone, e-mail e formacao publicados no site.",
        "identidade": "Wix com banner/Instagram e poucos elementos proprietarios.",
        "imagens": "Banner e identidade atual podem ser aproveitados.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Rio de Janeiro - RJ",
        "nome": "Beatriz Sa Orgal",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5521997320283",
        "site": "https://conexaotcc.wixsite.com/conexaotcc/rio-de-janeiro-cidades",
        "contato": "Telefone publico no site: (21) 99732-0283; e-mail beatrizsapsi@gmail.com.",
        "motivo": "Psicologa TCC com CRP e canais publicos, mas aparece em pagina Wix diretoria com baixa autoridade individual.",
        "diagnostico": "Presenca digital depende de listagem em Wix, sem landing individual forte, CTA visual ou identidade propria.",
        "conteudo": "Beatriz Sa Orgal, CRP 05/47591, psicologa TCC para criancas, adolescentes, adultos e idosos em Jacarepagua/Recreio.",
        "servicos": "TCC; ansiedade; depressao; regulacao emocional; autoestima; relacionamentos.",
        "prova": "CRP, telefone, e-mail, Facebook e Instagram publicados.",
        "identidade": "Listagem basica com foto pequena e texto padronizado.",
        "imagens": "Foto/listagem e redes sociais podem orientar pagina propria futura.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Francisco Beltrao - PR",
        "nome": "Emanuelle Martini",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5546999036530",
        "site": "https://psicoemanuelle.wixsite.com/website",
        "contato": "Telefone publico no site: (46) 99903-6530.",
        "motivo": "Psicologa mulher com telefone, CRP e especializacoes, mas site Wix simples e fora de padrao moderno.",
        "diagnostico": "Pagina Wix curta, CTA basico e pouca prova social; redesign pode destacar neuroaprendizagem e agendamento.",
        "conteudo": "Psicologa Emanuelle Martini, especializacoes em Neuroaprendizagem e Psicologia Comportamental e Cognitiva, mestranda em Educacao pela Unioeste.",
        "servicos": "Psicoterapia; psicologia comportamental e cognitiva; neuroaprendizagem.",
        "prova": "Telefone, formacao e especializacoes publicados no site.",
        "identidade": "Wix simples com foto de WhatsApp e poucos elementos de marca.",
        "imagens": "Foto/elementos do site podem ser usados como referencia.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Londrina - PR",
        "nome": "Gleicieni Quiel",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5543999800134",
        "site": "https://psigleiciquiel.wixsite.com/consultorio",
        "contato": "WhatsApp publico no site: (43) 99980-0134.",
        "motivo": "Psicologa em Londrina com recorte para mulheres, WhatsApp e CRP, mas site Wix antigo.",
        "diagnostico": "Site de 2018 com banner Wix, template antigo, CTA e prova social pouco explorados.",
        "conteudo": "Gleicieni Quiel, CRP 08/23494, graduada pela UEL, pos-graduada em Saude Mental; atendimento para mulheres.",
        "servicos": "Terapia online; atendimento para mulheres; ansiedade; independencia emocional; inseguranca e autoconfianca.",
        "prova": "CRP, WhatsApp e foco de atendimento publicados no site.",
        "identidade": "Wix antigo com foto e poucos elementos visuais.",
        "imagens": "Foto profissional e imagens da pagina podem ser reaproveitadas.",
    },
    {
        "nicho": "Psicologa",
        "regiao": "Sorocaba - SP",
        "nome": "Sara Possamai",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5515997777477",
        "site": "https://luispossamai.wixsite.com/15-9-9777-7477/about",
        "contato": "WhatsApp publico no site/blog: (15) 99777-7477.",
        "motivo": "Psicologa clinica em Sorocaba com telefone, mas site Wix antigo e URL improvisada com numero.",
        "diagnostico": "URL pouco profissional, blog antigo, textos longos e baixa clareza de conversao.",
        "conteudo": "Sara Possamai, psicologia clinica e psicoterapia, atende transtornos psiquiatricos, luto, conflitos pessoais e carreira.",
        "servicos": "Psicologia clinica; psicoterapia; transtornos psiquiatricos; luto; conflitos conjugais/familiares; carreira.",
        "prova": "WhatsApp e artigos/workshops publicados no site.",
        "identidade": "Wix antigo com blog e baixa identidade visual.",
        "imagens": "Fotos/artigos do site podem orientar nova pagina.",
    },
    {
        "nicho": "Nutricionista",
        "regiao": "Sao Paulo - SP",
        "nome": "Rita C. S. Fernandes",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5511976197929",
        "site": "https://ritafernandesnutri.my.canva.site/",
        "contato": "WhatsApp publico no site: 11976197929.",
        "motivo": "Nutricionista com CRN e WhatsApp, mas site Canva simples.",
        "diagnostico": "Pagina Canva de consultorio online, pouca hierarquia e baixo refinamento de autoridade.",
        "conteudo": "Nutricionista Rita C. S. Fernandes, CRN-3 56632, atendimentos nutricionais online e agendamento online.",
        "servicos": "Atendimento nutricional online; consultorio online; agendamento.",
        "prova": "CRN, WhatsApp e link de agendamento publicados no site.",
        "identidade": "Canva com visual basico e pouca marca proprietaria.",
        "imagens": "Elementos do Canva podem orientar redesign.",
    },
    {
        "nicho": "Nutricionista",
        "regiao": "Belo Horizonte - MG",
        "nome": "Gabrielle Navarro",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5531988381231",
        "site": "https://gabriellenavarro8.wixsite.com/nutricionista",
        "contato": "Telefone publico no site: (31) 98838-1231; e-mail gabrielle.navarro@hotmail.com.",
        "motivo": "Nutricionista materno-infantil com CRN e telefone, mas site Wix antigo.",
        "diagnostico": "Dominio Wix, menu simples e design datado; redesign pode valorizar nicho materno-infantil e CTA de agenda.",
        "conteudo": "Gabrielle Navarro Nutricionista Materno Infantil, CRN-9: 1888-7, Belo Horizonte, com consultas para tentantes, gestantes, lactantes, introducao alimentar e infancia.",
        "servicos": "Tentantes; gestantes; lactantes; introducao alimentar; infancia; e-book.",
        "prova": "CRN, telefone, e-mail e especialidade publicados no site.",
        "identidade": "Wix com imagens materno-infantis e estrutura basica.",
        "imagens": "Imagens e materiais do e-book podem ser aproveitados.",
    },
    {
        "nicho": "Nutricionista",
        "regiao": "Sorocaba - SP",
        "nome": "Lais Hess",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5515997639751",
        "site": "https://nutricionistalaishess.my.canva.site/",
        "contato": "Telefone/WhatsApp publico no site: (15) 99763-9751; e-mail nutricionistalaishess@gmail.com.",
        "motivo": "Nutricionista mulher com WhatsApp, mas presenca em Canva e baixa autoridade visual.",
        "diagnostico": "Site Canva de pagina unica, CTA simples e pouca estrutura para conversao e prova social.",
        "conteudo": "Nutricionista Lais Hess; suporte via WhatsApp diretamente com a profissional; e-mail publicado.",
        "servicos": "Atendimento nutricional; suporte via WhatsApp.",
        "prova": "Telefone/WhatsApp e e-mail publicados no site.",
        "identidade": "Canva simples com visual de landing rapida.",
        "imagens": "Elementos e imagens do Canva podem orientar redesign.",
    },
    {
        "nicho": "Nutricionista",
        "regiao": "Sao Paulo - SP",
        "nome": "Erica Arraes",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5511911409824",
        "site": "https://nutriericaarraes.wixsite.com/nutri",
        "contato": "WhatsApp publico no site: (11) 91140-9824; e-mail nutriericaarraes@gmail.com.",
        "motivo": "Nutricionista com CRN, consultorio e WhatsApp, mas site Wix ainda simples.",
        "diagnostico": "Pagina Wix com conteudo direto, porem visual basico, CTA pouco destacado e pouca prova social.",
        "conteudo": "Nutricionista Erica Arraes, CRN/SP 71033, consultorio na R. Manuel da Nobrega, 354 - Paraiso, Sao Paulo.",
        "servicos": "Atendimento nutricional particular; reembolso por convenios.",
        "prova": "CRN, endereco, e-mail e WhatsApp publicados no site.",
        "identidade": "Wix com visual simples e pouca diferenciacao.",
        "imagens": "Imagens e identidade atual podem orientar redesign.",
    },
    {
        "nicho": "Nutricionista",
        "regiao": "Sao Paulo - SP",
        "nome": "Lilian Vessoni Canhetti",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+551141064534",
        "site": "https://lilianvc.wixsite.com/nutricao",
        "contato": "Telefone publico no site: (11) 4106-4534.",
        "motivo": "Nutricionista clinica com endereco e telefone, mas site Wix antigo de 2012.",
        "diagnostico": "Layout muito datado, dominio Wix, CTA fraco e ausencia de prova social moderna.",
        "conteudo": "Nutricionista clinica Lilian Vessoni Canhetti, endereco Rua Coronel Lisboa, 735, Vila Clementino, Sao Paulo.",
        "servicos": "Nutricionista clinica; atendimento nutricional.",
        "prova": "Telefone, endereco e historico de site desde 2012 publicados.",
        "identidade": "Wix antigo, texto grande e baixa modernidade visual.",
        "imagens": "Imagens simples do site podem guiar antes/depois.",
    },
    {
        "nicho": "Fisioterapeuta",
        "regiao": "Sorocaba - SP",
        "nome": "Raphaela Souza",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5515988352165",
        "site": "https://raphaela9souza.wixsite.com/raphaelasouza",
        "contato": "Telefone publico no site: (15) 98835-2165; e-mail estudioraphaela@gmail.com.",
        "motivo": "Fisioterapeuta mulher com CREFITO e depoimento, mas site Wix com layout antigo.",
        "diagnostico": "Banner Wix, textos extensos, CTA disperso e prova social pouco valorizada.",
        "conteudo": "Raphaela Souza, CREFITO 3 184078-F, fisioterapeuta em Sorocaba, Pilates, LPF, atendimento para adultos, idosos e gestantes.",
        "servicos": "Pilates; LPF; fisioterapia baseada em exercicio; atendimento para adultos, idosos e gestantes.",
        "prova": "CREFITO, endereco, telefone, e-mail e depoimento publicados no site.",
        "identidade": "Wix com imagens de pilates e foto pessoal, mas pouco premium.",
        "imagens": "Fotos de servicos e depoimento podem ser reaproveitados.",
    },
    {
        "nicho": "Fisioterapeuta",
        "regiao": "Curitiba - PR",
        "nome": "Simone Lara",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5541999915703",
        "site": "https://simonelarafisio.wixsite.com/fisioterapia",
        "contato": "Telefone/WhatsApp publico no site: (41) 99991-5703; e-mail simonelarab@hotmail.com.",
        "motivo": "Clinica/fisioterapeuta em Curitiba com WhatsApp e endereco, mas site Wix antigo.",
        "diagnostico": "Site com cards simples, imagens genericas e CTA pouco forte; redesign pode organizar servicos e agenda.",
        "conteudo": "Clinica de fisioterapia Simone Lara, Rua Jose Brum 170, Xaxim, Curitiba; atendimento ortopedia, neurologica, drenagem, geriatria.",
        "servicos": "Fisioterapia ortopedica; neurologica; drenagem linfatica; geriatria; atendimento ambulatorial/domiciliar.",
        "prova": "Endereco, telefone/WhatsApp e e-mail publicados no site.",
        "identidade": "Wix com imagens genericas e estrutura simples.",
        "imagens": "Imagens de servicos e identidade atual podem orientar redesign.",
    },
    {
        "nicho": "Fisioterapeuta",
        "regiao": "Sao Paulo - SP",
        "nome": "Eliana Tessitore",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5511964560777",
        "site": "https://elianatessitore.wixsite.com/cuide-sebem",
        "contato": "Telefone publico no site: 11 96456-0777; e-mail elianatessitore@gmail.com.",
        "motivo": "Fisioterapeuta com CREFITO e experiencia, mas site Wix antigo e pouco comercial.",
        "diagnostico": "Site de 2016 em Wix, com contato no rodape e baixa hierarquia para conversao.",
        "conteudo": "Dra. Eliana Tessitore, fisioterapeuta CREFITO 3650F, mestre pela EEUSP/IPq, RPG Souchard, cursos e consultorias.",
        "servicos": "Fisioterapia; RPG Souchard; cursos; palestras; workshops; consultorias.",
        "prova": "CREFITO, e-mail, telefone e formacao publicados no site.",
        "identidade": "Wix antigo com linguagem institucional e visual simples.",
        "imagens": "Fotos e elementos atuais podem orientar pagina moderna.",
    },
    {
        "nicho": "Fisioterapeuta",
        "regiao": "Sao Paulo - SP",
        "nome": "Silvia Yukie R de Almeida",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5511962288033",
        "site": "https://silviayukiefisio.wixsite.com/yukie",
        "contato": "Telefone publico no site: 11 96228-8033.",
        "motivo": "Fisioterapeuta mulher com CREFITO e dois enderecos, mas site Wix antigo.",
        "diagnostico": "Site com banner Wix, menu longo, informacao de localizacao escondida e CTA fraco.",
        "conteudo": "Silvia Yukie R. de Almeida, fisioterapeuta CREFITO 3/66979-F, RPG Souchard e acupuntura em Cambuci/Mirandopolis.",
        "servicos": "Fisioterapia; RPG Metodo Souchard; acupuntura.",
        "prova": "CREFITO, telefone e enderecos publicados no site.",
        "identidade": "Wix antigo com layout de menu e conteudo datado.",
        "imagens": "Imagens/elementos do site podem orientar redesign.",
    },
    {
        "nicho": "Fisioterapeuta",
        "regiao": "Florianopolis - SC",
        "nome": "Ana Paula Ruschel",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5548998610922",
        "site": "https://anaruschelfisio.wixsite.com/site",
        "contato": "Telefone publico no site: (48) 99861-0922; e-mail anapruschel@yahoo.com.br.",
        "motivo": "Fisioterapeuta mulher em Florianopolis com telefone, mas site Wix antigo.",
        "diagnostico": "Pagina Wix simples, conteudo de microfisioterapia em texto longo e CTA pouco destacado.",
        "conteudo": "Dra. Ana Paula Ruschel, fisioterapeuta formada em 2006, microfisioterapia e atendimento em Florianopolis.",
        "servicos": "Microfisioterapia; fisioterapia domiciliar; tratamento pos-covid; terapia manual.",
        "prova": "Telefone, e-mail, cidade e historico profissional publicados.",
        "identidade": "Wix com imagens de fisioterapia e foto, pouca marca.",
        "imagens": "Fotos da profissional e imagens de fisioterapia podem ser aproveitadas.",
    },
    {
        "nicho": "Fisioterapeuta",
        "regiao": "Florianopolis - SC",
        "nome": "Jacira Guesser Kruger",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+554896036126",
        "site": "https://jaciraiaia.wixsite.com/fisioterapiacme",
        "contato": "Telefone publico no site: (48) 9603-6126; e-mail jacira.iaia@gmail.com.",
        "motivo": "Fisioterapeuta mulher com metodo CME e endereco, mas site Wix antigo.",
        "diagnostico": "Pagina Wix com layout datado, telefone antigo sem nono digito e CTA basico; redesign pode reforcar especializacao e agenda.",
        "conteudo": "Jacira Guesser Kruger, fisioterapeuta CME nivel 3, atua com bebes e criancas desde 1996, atendimento no Itacorubi, Florianopolis.",
        "servicos": "Fisioterapia Metodo Cuevas Medek Exercises; atendimento a bebes e criancas; programas por periodo.",
        "prova": "Conselho 22260-10, telefone, endereco e e-mail publicados.",
        "identidade": "Wix antigo, visual simples e imagens infantis.",
        "imagens": "Logo/imagens do metodo e fotos podem orientar redesign.",
    },
    {
        "nicho": "Terapeuta",
        "regiao": "Petropolis - RJ",
        "nome": "Espaco Gaia - Saude, Arte & Pensamento",
        "nota": "nao capturada",
        "avaliacoes": "nao capturadas",
        "phone": "+5524988173244",
        "site": "https://chehabpsi.wixsite.com/espacogaia/copia-atendimentos",
        "contato": "WhatsApp publico no site: (24) 98817-3244; e-mail espacogaia610@hotmail.com.",
        "motivo": "Espaco com psicologa clinica/terapeuta mulher em Petropolis, mas site Wix antigo e pagina interna pouco clara.",
        "diagnostico": "Pagina interna Wix, estrutura antiga, informacao espalhada e CTA pouco proeminente.",
        "conteudo": "Espaco Gaia em Petropolis, psicologa clinica e terapeuta pos-graduada em Terapia Atraves do Movimento; contato e sala 610 publicados.",
        "servicos": "Psicologia clinica; terapia atraves do movimento; corpo, meditacao, danca e movimento.",
        "prova": "WhatsApp, e-mail, endereco e descricao profissional publicados.",
        "identidade": "Wix antigo com linguagem artistica e pouca clareza comercial.",
        "imagens": "Imagens de arte/corpo e identidade local podem guiar redesign.",
    },
]


def slugify(value: str) -> str:
    table = str.maketrans(
        "áàãâäéèêëíìîïóòõôöúùûüçñÁÀÃÂÄÉÈÊËÍÌÎÏÓÒÕÔÖÚÙÛÜÇÑ",
        "aaaaaeeeeiiiiooooouuuucnAAAAAEEEEIIIIOOOOOUUUUCN",
    )
    cleaned = value.translate(table).lower()
    out = []
    last_dash = False
    for ch in cleaned:
        if ch.isalnum():
            out.append(ch)
            last_dash = False
        elif not last_dash:
            out.append("-")
            last_dash = True
    return "".join(out).strip("-")


def wa_link(phone: str) -> str:
    return "https://wa.me/" + "".join(ch for ch in phone if ch.isdigit())


def block_1(lead: dict) -> str:
    return (
        "Olá! Tudo bem?\n\n"
        "Aqui é o Leonardo Brasil. Encontrei o seu site e preparei, sem compromisso, "
        "uma nova versão com visual mais moderno e pronta para celular.\n\n"
        "Quer que eu te envie o link para ver como ficou?\n\n"
        "Me envie uma mensagem no WhatsApp 51 99256-8861 que eu te mando o link:\n"
        "https://wa.me/5551992568861"
    )


def block_2(lead: dict) -> str:
    return (
        f"Proposta posterior para {lead['nome']}: redesign individual da pagina por R$497, "
        "com hospedagem/manutencao por R$37,90 mensais quando fizer sentido apos aprovacao humana."
    )


def make_row(lead: dict, rank: int) -> dict:
    slug = f"{slugify(lead['nome'])}-{slugify(lead['regiao'].split('-')[0])}"
    maps_url = "https://www.google.com/maps/search/" + quote(f"{lead['nome']} {lead['regiao']}")
    base = PRINT_DIR / slug
    return {
        "data": DATE,
        "rank": rank,
        "nicho": lead["nicho"],
        "regiao": lead["regiao"],
        "nome": lead["nome"],
        "nota": lead["nota"],
        "avaliacoes": lead["avaliacoes"],
        "contato": lead["contato"],
        "WhatsApp/telefone normalizado": lead["phone"],
        "site atual": lead["site"],
        "link Maps": maps_url,
        "motivo da abordagem": lead["motivo"],
        "diagnostico do site atual": lead["diagnostico"],
        "conteudo real extraido": lead["conteudo"],
        "servicos reais identificados": lead["servicos"],
        "prova social real": lead["prova"],
        "identidade visual observada": lead["identidade"],
        "imagens/logo aproveitaveis": lead["imagens"],
        "slug sugerido": slug,
        "status redesign": "Aguardando redesign individual",
        "URL curta gerada": "",
        "print Maps": str(base) + "-maps.png",
        "print site atual desktop": str(base) + "-desktop.png",
        "print site atual mobile": str(base) + "-mobile.png",
        "link wa.me preliminar": wa_link(lead["phone"]),
        "status": "LEAD_QUALIFICADO_AGUARDANDO_REDESIGN_INDIVIDUAL",
        "observacoes": "Lead novo apos dedupe local por dominio, telefone e nome; nao houve envio automatico nem criacao de redesign.",
        "Bloco 1 (Link)": block_1(lead),
        "Bloco 2 (Proposta)": block_2(lead),
        "Toque 1 (2 dias)": "Follow-up manual 2 dias apos primeiro contato, sem automacao de envio.",
        "Toque 2 (5 dias)": "Follow-up manual 5 dias depois, somente se houver canal adequado e confirmacao humana.",
        "Toque 3 (10 dias)": "Ultimo toque manual em 10 dias, respeitando opt-out e sem insistencia.",
    }


def write_outputs():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PRINT_DIR.mkdir(parents=True, exist_ok=True)
    rows = [make_row(lead, idx + 1) for idx, lead in enumerate(LEADS)]

    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)

    JSON_PATH.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = DATE
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row[col] for col in COLUMNS])

    ws.freeze_panes = "K2"
    ws.auto_filter.ref = f"A1:AF{ws.max_row}"

    header_fill = PatternFill("solid", fgColor="111111")
    critical_fill = PatternFill("solid", fgColor="FFD700")
    white_font = Font(color="FFFFFF", bold=True)
    black_font = Font(color="111111", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    critical_cols = {9, 10, 21, 25, 26}
    for cell in ws[1]:
        cell.fill = critical_fill if cell.column in critical_cols else header_fill
        cell.font = black_font if cell.column in critical_cols else white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        row[8].number_format = "@"
        row[20].value = None

    widths = {
        "A": 12,
        "B": 8,
        "C": 22,
        "D": 24,
        "E": 30,
        "F": 16,
        "G": 16,
        "H": 42,
        "I": 24,
        "J": 46,
        "K": 50,
        "L": 56,
        "M": 62,
        "N": 72,
        "O": 58,
        "P": 50,
        "Q": 50,
        "R": 48,
        "S": 38,
        "T": 30,
        "U": 20,
        "V": 52,
        "W": 52,
        "X": 52,
        "Y": 34,
        "Z": 52,
        "AA": 52,
        "AB": 68,
        "AC": 58,
        "AD": 48,
        "AE": 52,
        "AF": 48,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 42
    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 150

    dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_VALUES) + '"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"Z2:Z{ws.max_row}")

    status_colors = {
        "LEAD_QUALIFICADO_AGUARDANDO_REDESIGN_INDIVIDUAL": "D1EFD1",
        "SEM_SITE_NAO_CONTA": "E5E5E5",
        "SITE_BOM_DESCARTADO": "CCE5FF",
        "NICHO_EXCLUIDO_NAO_CONTA": "E0D1F4",
        "CONTATO_INSUFICIENTE": "FFF4B3",
        "META_NAO_BATIDA": "FFC6C6",
    }
    for status, color in status_colors.items():
        ws.conditional_formatting.add(
            f"A2:AF{ws.max_row}",
            FormulaRule(formula=[f'$Z2="{status}"'], fill=PatternFill("solid", fgColor=color)),
        )
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                if cell.column != 26:
                    cell.fill = PatternFill("solid", fgColor="FAFAFA")

    summary = wb.create_sheet("Resumo")
    summary_rows = [
        ["item", "valor"],
        ["data", DATE],
        ["leads qualificados", len(rows)],
        ["status principal", "LEAD_QUALIFICADO_AGUARDANDO_REDESIGN_INDIVIDUAL"],
        ["URL curta gerada", "em branco por regra da automacao"],
        ["envio automatico", "nao realizado"],
        ["redesign/build/publicacao", "nao realizado"],
        ["observacao", "Aba local criada; importar XLSX no Drive para virar Google Sheets nativo."],
    ]
    for item in summary_rows:
        summary.append(item)
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 86
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    wb.save(XLSX_PATH)
    print(CSV_PATH)
    print(XLSX_PATH)
    print(JSON_PATH)


if __name__ == "__main__":
    write_outputs()
